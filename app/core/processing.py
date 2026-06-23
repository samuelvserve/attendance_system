import re
import os
import tempfile
import requests
import numpy as np
import pandas as pd
import pytz
import warnings
from datetime import datetime, timedelta
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side

warnings.filterwarnings("ignore")

# ============================================================================
# LIVE LOGGING CALLBACK
# All processing functions accept an optional `log_cb(msg, type)` callable.
# When omitted a silent no-op is used so the functions work standalone too.
# ============================================================================

def _noop_log(msg, log_type="info"):
    pass


# ============================================================================
# UTILITY / HELPER FUNCTIONS  (ported 1-to-1 from biometric_timechamp_report.py)
# ============================================================================

def extract_in_out_events(punch_records):
    """Extract IN/OUT times from a punch-record string like '09:05:in,13:00:out,...'"""
    matches = re.findall(r'(\d{2}:\d{2}):(in|out)', punch_records)
    in_events  = [t for t, s in matches if s == 'in']
    out_events = [t for t, s in matches if s == 'out']
    in_out_dict = {}
    for i, t in enumerate(in_events):
        in_out_dict[f'IN {i+1}']  = t if t else pd.NaT
    for i, t in enumerate(out_events):
        in_out_dict[f'OUT {i+1}'] = t if t else pd.NaT
    return pd.Series(in_out_dict)


def adjust_timestamp(time, date_str):
    """Prefix a HH:MM string with its date; push past-midnight times to next day."""
    if isinstance(time, str):
        timestamp    = f"{date_str} {time}"
        timestamp_dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M")
        if timestamp_dt.time() < datetime.strptime("07:00", "%H:%M").time():
            timestamp_dt += timedelta(days=1)
        return timestamp_dt.strftime("%Y-%m-%d %H:%M")
    return pd.NaT


def find_last_non_nan(row):
    for value in row:
        if not pd.isna(value):
            return value
    return pd.NaT


def count_in_out(row):
    in_cols  = [c for c in row.index if 'IN '  in c]
    out_cols = [c for c in row.index if 'OUT ' in c]
    return (
        sum(1 for c in in_cols  if pd.notna(row[c])),
        sum(1 for c in out_cols if pd.notna(row[c])),
    )


def work_time(val):
    """Total minutes between First_IN and Last_OUT."""
    first_in  = pd.to_datetime(val['First_IN'])
    last_out  = pd.to_datetime(val['Last_OUT'])
    if pd.isnull(first_in) or pd.isnull(last_out):
        return pd.NaT
    return int((last_out - first_in).total_seconds() // 60)


def number_to_hr(time_str):
    """Convert a pd.Timedelta to a human-readable string like '7 hrs 30 mins'."""
    if pd.isna(time_str):
        return pd.NaT
    total_seconds    = time_str.total_seconds()
    minutes          = total_seconds // 60
    hours            = int(minutes // 60)
    remaining_mins   = int(minutes % 60)
    remaining_secs   = int(total_seconds % 60)
    if hours and remaining_mins:
        return f"{hours} hrs {remaining_mins} mins"
    elif hours and remaining_secs:
        return f"{hours} hrs {remaining_secs} secs"
    elif remaining_mins and remaining_secs:
        return f"{remaining_mins} mins {remaining_secs} secs"
    elif hours:
        return f"{hours} hrs"
    elif remaining_mins:
        return f"{remaining_mins} mins"
    elif remaining_secs:
        return f"{remaining_secs} secs"
    return pd.NaT


def clean_time_string(time_str):
    """Parse a time string / timedelta to pd.Timedelta, returning NaT on failure."""
    try:
        if pd.isna(time_str) or time_str == '':
            return pd.NaT
        return pd.to_timedelta(str(time_str))
    except Exception:
        return pd.NaT


def parse_time_string(x):
    """Parse a human-readable duration like '7 hrs 30 mins' to pd.Timedelta."""
    if not isinstance(x, str):
        return pd.Timedelta(0)
    parts = x.split()
    hours = minutes = seconds = 0
    for i in range(0, len(parts), 2):
        value = int(parts[i])
        unit  = parts[i + 1] if i + 1 < len(parts) else 'secs'
        if 'hr'  in unit: hours   = value
        elif 'min' in unit: minutes = value
        elif 'sec' in unit: seconds = value
    return pd.Timedelta(hours=hours, minutes=minutes, seconds=seconds)


def OutPunch_status(val):
    if val['IN_Time Count'] != val['OUT_Time Count']:
        return 'User No OutPunch'


def attendance_determine_status(team, worked_hours):
    hours = float(worked_hours) if not pd.isna(worked_hours) else 0
    if team in ('Rollkall', 'TE'):
        if hours == 0:               return 'Absent'
        elif 3.75 <= hours <= 7.24:  return 'Half day'
        elif hours >= 7.25:          return 'Full day'
        else:                        return 'Absent'
    else:
        if hours == 0:               return 'Absent'
        elif 3.75 <= hours < 7.74:   return 'Half day'
        elif hours >= 7.75:          return 'Full day'
        else:                        return 'Absent'


def below_7_5_hours(team, hours):
    if pd.isna(hours):
        return pd.NA
    hours = float(hours)
    if team in ('Rollkall', 'TE'):
        if 0 <= hours < 3.75:        return 'Worked below 4 hrs'
        elif 3.75 <= hours < 7.25:   return 'Worked below 8 hrs'
    else:
        if 0 <= hours < 3.75:        return 'Worked below 4 hrs'
        elif 3.75 <= hours < 7.75:   return 'Worked below 8 hrs'


def is_break_greater_than_one_hour(team, break_hours):
    if pd.isna(break_hours):
        return pd.NA
    hours = float(break_hours)
    if team in ('Rollkall', 'TE'):
        return 'Break more than 1 hr 30 mins' if hours >= 1.59 else pd.NaT
    else:
        return 'Break more than 1 hr' if hours > 1.17 else pd.NaT


def convert_shift_format(shift):
    if pd.isna(shift):
        return pd.NaT
    shift = str(shift).strip()
    if shift == 'GS': return '09:00 IST - 18:00 IST'
    if shift == 'NS': return '22:00 IST - 06:00 IST'
    if 'IST' in shift and ' - ' in shift: return shift
    return shift


def get_shift_start(shift_str):
    if isinstance(shift_str, str):
        start_time_str = shift_str.split(' - ')[0].strip().replace('IST', '').strip()
        try:
            return pd.to_datetime(start_time_str, format='%H:%M').time()
        except ValueError:
            return pd.NaT
    return pd.NaT


def calculate_late_login(in_time_str, shift_start_time):
    """Return a human-readable late-login string, 'Early Login', or 'No Delay'."""
    if pd.isna(in_time_str) or pd.isna(shift_start_time):
        return pd.NaT
    try:
        in_time  = pd.to_datetime(in_time_str).time()
        today    = datetime.today().date()
        # Grace window: shift start + 30 min (matches original script)
        shift_dt = datetime.combine(today, shift_start_time) + timedelta(minutes=30)
        login_dt = datetime.combine(today, in_time)
        # Overnight shift
        if shift_start_time.hour >= 18 and in_time.hour < 12:
            login_dt += timedelta(days=1)
        late_time = login_dt - shift_dt
        if late_time < timedelta(0):
            return "Early Login"
        if late_time <= timedelta(minutes=5):
            return "No Delay"
        total_mins = int(late_time.total_seconds() // 60)
        hrs, mins  = divmod(total_mins, 60)
        if hrs and mins: return f"{hrs} hrs {mins} mins late"
        elif hrs:        return f"{hrs} hrs late"
        else:            return f"{mins} mins late"
    except Exception:
        return pd.NaT


def calculate_early_logout(out_time_str, shift, grace_min=5):
    """Return early-logout string or 'No Early Logout'. Handles full datetime strings."""
    if pd.isna(out_time_str) or pd.isna(shift):
        return pd.NaT
    try:
        start_str, end_str = [re.sub(r'[A-Z]+', '', x).strip() for x in str(shift).split(' - ')]
        shift_start = datetime.strptime(start_str, "%H:%M")
        shift_end   = datetime.strptime(end_str,   "%H:%M")
        # Handle both "HH:MM" and full datetime strings
        try:
            out_time = datetime.strptime(str(out_time_str), "%H:%M")
        except ValueError:
            parsed   = pd.to_datetime(out_time_str)
            out_time = datetime.strptime(parsed.strftime("%H:%M"), "%H:%M")
    except Exception:
        return pd.NaT

    today       = datetime.today().date()
    shift_start = datetime.combine(today, shift_start.time())
    shift_end   = datetime.combine(today, shift_end.time())
    out_time    = datetime.combine(today, out_time.time())

    if shift_end <= shift_start:          # overnight
        shift_end += timedelta(days=1)
        if out_time.hour < 12:
            out_time += timedelta(days=1)

    diff = shift_end - out_time
    if diff <= timedelta(minutes=grace_min):
        return "No Early Logout"

    hours, remainder = divmod(int(diff.total_seconds()), 3600)
    minutes = remainder // 60
    if hours and minutes: return f"{hours} hrs {minutes} mins early"
    elif hours:           return f"{hours} hrs early"
    return f"{minutes} mins early"


# ── Styling helpers ──────────────────────────────────────────────────────────

def highlight_status(val, status):
    if isinstance(status, str):
        return 'background-color: red'
    return ''


def break_highlight_status(row):
    if isinstance(row.get('Break Hours Status'), str):
        return ['background-color: red' if col == 'Break Hours' else '' for col in row.index]
    return [''] * len(row)


def Attendance_highlight_status(value):
    if not isinstance(value, str):
        return ''
    if 'Full day'  in value: return 'background-color: #89fe79; color: black'
    if 'Half day'  in value: return 'background-color: #ffb3b3; color: black'
    if 'Absent'    in value: return 'background-color: #fe7d79; color: black'
    if 'WeekOff'   in value: return 'background-color: #ffd700; color: black'
    return ''


def Date_highlight_status(val):
    try:
        date = datetime.strptime(str(val), '%Y-%m-%d')
    except Exception:
        return ''
    if date.weekday() == 6:   # Sunday
        return 'background-color: #FFD700; color: black'
    if date.weekday() == 5:   # Saturday — only 1st & 3rd week
        week_of_month = (date.day - 1) // 7 + 1
        if week_of_month in [1, 3]:
            return 'background-color: #FFD700; color: black'
    return ''


# ── Misc helpers ─────────────────────────────────────────────────────────────

def normalize_shift(value):
    if pd.isna(value):
        return value
    s = re.sub(r'\s+', ' ', str(value).strip().upper())
    if s == "GS":   return "09:00 IST - 18:00 IST"
    if s == "NS":   return "22:00 IST - 06:00 IST"
    if s == "XMEK": return "XMEK"
    m = re.search(r'(\d{1,2})\.(\d{2}).*?TO\s*(\d{1,2})\.(\d{2})', s)
    if m:
        sh, sm, eh, em = m.groups()
        return f"{int(sh):02d}:{sm} IST - {int(eh):02d}:{em} IST"
    return s


def convert_to_cst(time_value, team):
    """Convert IST shift/time to US Central (DST-aware). Only for Rollkall team."""
    if team != 'Rollkall' or pd.isna(time_value):
        return time_value
    ist_tz = pytz.timezone('Asia/Kolkata')
    cst_tz = pytz.timezone('US/Central')
    value  = str(time_value).strip()
    if ' - ' in value:
        value = value.replace('IST', '')
        start_str, end_str = value.split(' - ')
        start_cst = ist_tz.localize(datetime.strptime(start_str.strip(), "%H:%M")).astimezone(cst_tz)
        end_cst   = ist_tz.localize(datetime.strptime(end_str.strip(),   "%H:%M")).astimezone(cst_tz)
        return f"{start_cst.strftime('%H:%M CST')} - {end_cst.strftime('%H:%M CST')}"
    else:
        single_cst = ist_tz.localize(
            datetime.strptime(value.replace('IST', '').strip(), "%H:%M")
        ).astimezone(cst_tz)
        return single_cst.strftime("%H:%M")


def map_late_login_status(status):
    if pd.isna(status) or status in ['Early Login', 'No Delay']:
        return 0
    return 1


def hr_api_call(min_date, max_date):
    headers = {
        "Cookie": "humans_21909=1",
        "User-Agent": "curl/8.0",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.post(
        'https://vservehr.com/attendance_api.php',
        data={'from_date': min_date, 'to_date': max_date},
        headers=headers,
    )
    return response.json()['data']


def attendance_mismatch(attendance_row, hr_attendance_row, half_leaves):
    if hr_attendance_row == 'Present' and attendance_row in ['Half day', 'Absent', 'WeekOff']:
        return 'YES'
    elif hr_attendance_row in half_leaves and attendance_row in ['Absent', 'WeekOff']:
        return 'YES'
    elif hr_attendance_row == 'No Data':
        return 'YES'
    return pd.NaT


def _get_actual_shift_time(shift):
    """Shrink a shift window by 30 min on each side for actual expected hours."""
    if pd.isna(shift):
        return pd.NaT
    try:
        start, end    = str(shift).split(' - ')
        start_time    = datetime.strptime(start.strip(), '%H:%M IST')
        end_time      = datetime.strptime(end.strip(),   '%H:%M IST')
        actual_start  = (start_time + timedelta(minutes=30)).strftime('%H:%M')
        actual_end    = (end_time   - timedelta(minutes=30)).strftime('%H:%M')
        return f"{actual_start} IST - {actual_end} IST"
    except Exception:
        return pd.NaT


# ============================================================================
# FINAL ADJUSTMENTS  (HR API merge, attendance status, CST conversion, styling)
# ============================================================================

def final_adjustments(df, log_cb=None):
    if log_cb is None:
        log_cb = _noop_log

    log_cb("📊 Calculating attendance status (Full day / Half day / Absent)...", "info")
    df['Attendance'] = df.apply(
        lambda r: attendance_determine_status(r['Team/Department'], r['Worked Hours (Number)']), axis=1
    )

    log_cb("⏱️  Evaluating break hour violations...", "info")
    df['Break Hours Status']  = df.apply(
        lambda r: is_break_greater_than_one_hour(r['Team/Department'], r['Break Hours (Number)']), axis=1
    )
    df['Worked Hours Status'] = df.apply(
        lambda r: below_7_5_hours(r['Team/Department'], r['Worked Hours (Number)']), axis=1
    )

    log_cb("🕐 Converting timedeltas to human-readable hours...", "info")
    for col in ['Worked Hours', 'Idle Hours', 'Break Hours']:
        if col in df.columns:
            df[col] = df[col].apply(number_to_hr)

    # Clear status for flex/24-hr shifts
    flex_shifts = ['00:00 IST - 23:55 IST', '12:00 IST - 11:55 IST', '00:00 IST - 23:59 IST']
    df.loc[df['Shift Timing'].isin(flex_shifts), 'Actual Shift Time'] = df['Shift Timing']

    log_cb("🚪 Calculating early logout status...", "info")
    df['Early Logout Status'] = df.apply(
        lambda r: calculate_early_logout(r['Logout Time'], r['Actual Shift Time']), axis=1
    )
    df.loc[df['Shift Timing'].isin(flex_shifts), 'Late Login Status']    = pd.NaT
    df.loc[df['Shift Timing'].isin(flex_shifts), 'Early Logout Status']  = pd.NaT

    log_cb("📅 Marking weekend days as WeekOff...", "info")
    df['Date'] = pd.to_datetime(df['Date'], format='%Y-%m-%d', errors='coerce')
    df['Day']  = df['Date'].dt.strftime('%a')
    df.loc[(df['Day'].isin(['Sat', 'Sun'])) & (df['Worked Hours'].isnull()), 'Attendance'] = 'WeekOff'

    min_date = df['Date'].min().strftime('%Y-%m-%d')
    max_date = df['Date'].max().strftime('%Y-%m-%d')
    df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')

    log_cb(f"🌐 Fetching HR attendance data ({min_date} → {max_date})...", "info")
    try:
        data = hr_api_call(min_date, max_date)
        log_cb(f"   ↳ HR API returned {len(data)} records", "info")
    except Exception as e:
        raise ValueError(f"Error calling HR API: {e}")

    hr_df = pd.DataFrame(data).astype(str)
    hr_df.rename(columns={'emp_id': 'Employee ID', 'date': 'Date', 'attendance': 'HR Attendance'}, inplace=True)

    df['Employee ID']    = df['Employee ID'].astype(str).str.strip()
    hr_df['Employee ID'] = hr_df['Employee ID'].astype(str).str.strip()
    df['Date']           = pd.to_datetime(df['Date']).dt.date.astype(str)
    hr_df['Date']        = pd.to_datetime(hr_df['Date']).dt.date.astype(str)

    log_cb("🔗 Merging HR attendance with processed data...", "info")
    df = pd.merge(df, hr_df, how='left', on=['Employee ID', 'Date'])
    df['HR Attendance'].fillna('No Data', inplace=True)

    matched   = (df['HR Attendance'] != 'No Data').sum()
    unmatched = (df['HR Attendance'] == 'No Data').sum()
    log_cb(f"   ↳ HR matched: {matched} records  |  No HR data: {unmatched} records", "info")

    leave_shifts = [
        'Leave - Full Day', 'Absent - Full Day', 'Sick Leave - Full',
        'EL FD', 'Maternity Leave', 'Bereavement Leave', 'Week Off',
        'Paternity Leave', 'Adoption Leave', 'Paid Holiday', 'Comp Off',
    ]
    cols_to_null = [
        'Login Time', 'Logout Time', 'Late Login Status', 'Worked Hours', 'Worked Hours (Number)',
        'Idle Hours', 'Away Hours', 'Break Hours', 'Break Hours (Number)', 'Worked Hours Status',
        'Break Hours Status', 'No of Breaks (Biometric)',
        'No of Break Counts more than 4 Times (Biometric)', 'Comments',
    ]
    df.loc[df['HR Attendance'].isin(leave_shifts), cols_to_null] = np.nan

    half_leaves = ['Sick Leave - Half', 'EL HD', 'Leave - Half Day', 'Absent - Half Day']
    df.loc[df['HR Attendance'].isin(half_leaves) & (df['Worked Hours Status'] == 'Worked below 8 hrs'),
           'Worked Hours Status'] = np.nan
    df.loc[df['HR Attendance'].isin(half_leaves), 'Break Hours Status']   = np.nan
    df.loc[df['HR Attendance'].isin(half_leaves), 'Early Logout Status']  = np.nan
    df.loc[df['HR Attendance'].isin(half_leaves), 'Late Login Status']    = np.nan
    df.loc[(df['Worked Hours (Number)'].fillna(0) < 1) & (df['HR Attendance'] != 'Present'),
           'Late Login Status']   = pd.NaT
    df.loc[(df['Worked Hours (Number)'].fillna(0) < 1) & (df['HR Attendance'] != 'Present'),
           'Early Logout Status'] = pd.NaT

    log_cb("🔍 Computing attendance mismatch flags...", "info")
    df['Attendance Mismatch'] = df.apply(
        lambda r: attendance_mismatch(r['Attendance'], r['HR Attendance'], half_leaves), axis=1
    )
    mismatch_count = df['Attendance Mismatch'].eq('YES').sum()
    if mismatch_count:
        log_cb(f"   ↳ ⚠️  {mismatch_count} attendance mismatches detected", "warning")
    else:
        log_cb("   ↳ No attendance mismatches ✓", "info")

    # Final column selection — use only columns that exist (handles all 3 source types)
    base_cols = ['Date', 'Day', 'Employee ID', 'Employee Name', 'Team/Department',
                 'Source', 'Attendance', 'HR Attendance', 'Attendance Mismatch',
                 'Shift Timing', 'Actual Shift Time', 'Login Time', 'Logout Time',
                 'Late Login Status', 'Early Logout Status',
                 'Worked Hours', 'Worked Hours (Number)', 'Worked Hours Status',
                 'Break Hours', 'Break Hours (Number)', 'Break Hours Status',
                 'Idle Hours', 'Away Hours',
                 'No of Breaks (Biometric)', 'No of Break Counts more than 4 Times (Biometric)',
                 'Comments']
    df = df[[c for c in base_cols if c in df.columns]]

    log_cb("🌏 Applying IST → CST timezone conversion for Rollkall team...", "info")
    for col in ['Shift Timing', 'Actual Shift Time', 'Login Time', 'Logout Time']:
        if col in df.columns:
            df[col] = df.apply(lambda r: convert_to_cst(r[col], r['Team/Department']), axis=1)

    df['Attendance'].fillna('Absent', inplace=True)
    df = df.replace(0, pd.NaT)
    df['Employee Name'] = df['Employee Name'].replace(r'\s+', ' ', regex=True).str.strip()
    df = df.sort_values(by=['Employee Name', 'Date'])
    df.insert(0, 'SL No.', range(1, len(df) + 1))

    log_cb("🎨 Applying cell colour styles (attendance, breaks, weekends)...", "info")
    df_styled = df.style.apply(break_highlight_status, axis=1)
    df_styled = df_styled.map(Attendance_highlight_status, subset=['Attendance'])
    df_styled = df_styled.map(Date_highlight_status,       subset=['Date'])
    return df_styled


# ============================================================================
# BIOMETRIC PROCESSOR
# ============================================================================

def process_biometric_data(df, log_cb=None):
    if log_cb is None:
        log_cb = _noop_log
    try:
        log_cb(f"📋 Loaded {len(df):,} raw records from biometric file", "info")

        log_cb("📅 Parsing and normalising dates...", "info")
        df['Date'] = pd.to_datetime(df['Date'], format='%Y-%m-%d', dayfirst=True).dt.strftime('%Y-%m-%d')
        log_cb(f"   ↳ {df['Date'].nunique()} unique dates  ({df['Date'].min()}  →  {df['Date'].max()})", "info")

        log_cb("🔄 Normalising shift codes (GS/NS → IST times)...", "info")
        df['Shift'] = df['Shift'].apply(normalize_shift)
        for shift, cnt in df['Shift'].value_counts().items():
            log_cb(f"   ↳ {shift}: {cnt} records", "info")

        df     = df.reset_index()
        maindf = df.copy()

        punch_null = df['Punch Records'].isna().sum()
        df = df[df['Punch Records'].notnull()]
        log_cb(f"👆 Extracting punch events — {len(df):,} with punches, {punch_null} with no punch data", "info")

        in_out_df  = df['Punch Records'].apply(extract_in_out_events)
        final_df   = pd.concat([df, in_out_df], axis=1)

        in_columns  = sorted([c for c in final_df.columns if c.startswith('IN')],  key=lambda x: int(x[2:]))
        out_columns = sorted([c for c in final_df.columns if c.startswith('OUT')], key=lambda x: int(x[3:]))
        interleaved = [col for pair in zip(in_columns, out_columns) for col in pair]
        if len(in_columns) > len(out_columns):
            interleaved += in_columns[len(out_columns):]
        elif len(out_columns) > len(in_columns):
            interleaved += out_columns[len(in_columns):]

        log_cb(f"   ↳ Max punch depth: {len(in_columns)} IN / {len(out_columns)} OUT columns", "info")
        df = final_df[['index', 'Date', 'Shift'] + interleaved]

        log_cb("🌙 Adjusting timestamps for overnight / NS shifts...", "info")
        time_cols = [c for c in df.columns if "IN" in c or "OUT" in c]
        for idx, row in df.iterrows():
            for col in time_cols:
                df.at[idx, col] = adjust_timestamp(row[col], row['Date'])

        log_cb("⏱️  Calculating First-IN and Last-OUT per employee per day...", "info")
        in_filter = df.filter(like='IN')
        df['First_IN'] = in_filter.iloc[:, 0] if not in_filter.empty else pd.NaT
        out_filter = df.filter(like='OUT')
        if not out_filter.empty:
            df['Last_OUT'] = out_filter.iloc[:, ::-1].apply(find_last_non_nan, axis=1)
        else:
            df['Last_OUT'] = pd.NaT

        df[["IN_Time Count", "OUT_Time Count"]] = df.apply(count_in_out, axis=1).tolist()
        df.columns = df.columns.str.strip()

        missing_out = (df['IN_Time Count'] != df['OUT_Time Count']).sum()
        if missing_out:
            log_cb(f"⚠️  {missing_out} records have mismatched IN/OUT punches", "warning")
        else:
            log_cb("   ↳ All IN/OUT punches balanced ✓", "info")

        log_cb("🧮 Computing worked hours and break durations...", "info")
        for i in range(1, 30):
            in_col1  = f'IN {i}'
            out_col1 = f'OUT {i}'
            in_col2  = f'IN {i+1}'
            out_col2 = f'OUT {i+1}'
            if out_col1 in df.columns and in_col1 in df.columns:
                df[f'Work_{i}']  = (
                    pd.to_datetime(df[out_col1], format='%Y-%m-%d %H:%M', errors='coerce') -
                    pd.to_datetime(df[in_col1],  format='%Y-%m-%d %H:%M', errors='coerce')
                ).dt.seconds // 60
            # Break = IN{i+1} - OUT{i+1}: both columns must exist
            if in_col2 in df.columns and out_col2 in df.columns:
                df[f'Break_{i}'] = (
                    pd.to_datetime(df[in_col2],  format='%Y-%m-%d %H:%M', errors='coerce') -
                    pd.to_datetime(df[out_col2], format='%Y-%m-%d %H:%M', errors='coerce')
                ).dt.seconds // 60

        work_cols  = [c for c in df.columns if c.startswith('Work_')]
        break_cols = [c for c in df.columns if c.startswith('Break_')]

        df['Worked Hours'] = pd.to_timedelta(df[work_cols].sum(axis=1)  if work_cols  else 0, unit='m')
        df['Break Hours']  = pd.to_timedelta(df[break_cols].sum(axis=1) if break_cols else 0, unit='m')

        df['Comments'] = df.apply(OutPunch_status, axis=1)

        # Use total_seconds() (not .seconds) to handle shifts > 24 hrs correctly
        df['Worked Hours (Number)'] = df['Worked Hours'].apply(
            lambda x: round(pd.Timedelta(x).total_seconds() / 3600, 2) if x else pd.NaT)
        df['Break Hours (Number)']  = df['Break Hours'].apply(
            lambda x: round(pd.Timedelta(x).total_seconds() / 3600, 2) if x else pd.NaT)

        avg_work = df['Worked Hours (Number)'].dropna().mean()
        log_cb(f"   ↳ Average worked hours per record: {avg_work:.2f} hrs", "info")

        log_cb("🔗 Merging punch data with employee master records...", "info")
        df = pd.merge(maindf, df, how='left', on='index')
        df.rename(columns={'Team': 'Team Name'}, inplace=True)
        df['Source'] = 'Biometric'

        df['No of Breaks (Biometric)'] = df['IN_Time Count'] - 1
        high_breaks = df['No of Breaks (Biometric)'].gt(4).sum()
        if high_breaks:
            log_cb(f"☕ {high_breaks} records with more than 4 breaks detected", "warning")

        df['No of Break Counts more than 4 Times (Biometric)'] = (
            df['No of Breaks (Biometric)'].gt(4).map({True: 'Yes', False: pd.NA})
        )

        df['Shift Timing']    = df['Shift_x'].apply(convert_shift_format)
        df['Actual Shift Time'] = df['Shift Timing']
        df['Shift Start']     = df['Shift Timing'].apply(get_shift_start)

        log_cb("🕐 Calculating late login status...", "info")
        df['Late Login Status'] = df.apply(
            lambda r: calculate_late_login(r['First_IN'], r['Shift Start']), axis=1
        )
        late_count = df['Late Login Status'].apply(
            lambda x: 0 if pd.isna(x) or x in ('Early Login', 'No Delay') else 1
        ).sum()
        log_cb(f"   ↳ Late logins found: {late_count}", "info" if late_count == 0 else "warning")

        df['Login Time']  = pd.to_datetime(df['First_IN']).dt.strftime("%H:%M")
        df['Logout Time'] = pd.to_datetime(df['Last_OUT']).dt.strftime("%H:%M")
        df['Idle Hours']  = pd.NaT
        df['Away Hours']  = pd.NaT

        df.rename(columns={
            'Employee Code': 'Employee ID',
            'Team Name': 'Team/Department',
            'Date_x': 'Date',
        }, inplace=True)

        log_cb("🔧 Running final adjustments (HR API, status flags, CST conversion)...", "info")
        df = final_adjustments(df, log_cb=log_cb)

        employees = df.data['Employee ID'].nunique() if hasattr(df, 'data') else '?'
        log_cb(f"✅ Biometric processing complete — {employees} employees", "success")
        return df

    except Exception as e:
        raise Exception(f"Error processing biometric data: {str(e)}")


# ============================================================================
# TIMECHAMP PROCESSOR
# ============================================================================

def process_timechamp_data(df, log_cb=None):
    if log_cb is None:
        log_cb = _noop_log
    try:
        log_cb(f"📋 Loaded {len(df):,} raw records from Timechamp file", "info")

        log_cb("📅 Parsing Timechamp date format...", "info")
        try:
            df['Date'] = pd.to_datetime(df['Date'].str.strip(), format='%b-%d-%Y', errors='coerce').dt.strftime('%Y-%m-%d')
        except Exception:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce', format='mixed').dt.strftime('%Y-%m-%d')
        log_cb(f"   ↳ Date range: {df['Date'].min()}  →  {df['Date'].max()}  ({df['Date'].nunique()} unique dates)", "info")

        df['Source'] = 'Timechamp'
        df.rename(columns={'User': 'Employee Name'}, inplace=True)

        required_cols = ['Date', 'Employee Id', 'Employee Name', 'Team Name', 'Source', 'Shift',
                         'In Time', 'Out Time', 'Working Hours', 'Away Time', 'Idle Time', 'Break Time']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        df = df[required_cols]

        log_cb("⏱️  Converting time columns to timedelta...", "info")
        for col in ['Working Hours', 'Idle Time', 'Away Time', 'Break Time']:
            df[col] = df[col].apply(clean_time_string)

        log_cb("🕐 Extracting login / logout times...", "info")
        df['Login Time']  = pd.to_datetime(df['In Time'].astype(str),  format='mixed').dt.strftime('%H:%M')
        df['Logout Time'] = pd.to_datetime(df['Out Time'].astype(str), format='mixed').dt.strftime('%H:%M')

        # Use total_seconds() for correctness on shifts > 24 h
        df['Worked Hours (Number)'] = df['Working Hours'].apply(
            lambda x: round(pd.Timedelta(x).total_seconds() / 3600, 2) if x else pd.NaT)
        df['Break Hours (Number)']  = df['Away Time'].apply(
            lambda x: round(pd.Timedelta(x).total_seconds() / 3600, 2) if x else pd.NaT)
        df['Break Time'] = df['Break Time'].apply(
            lambda x: round(pd.Timedelta(x).total_seconds() / 3600, 2) if x else pd.NaT)

        avg_work   = df['Worked Hours (Number)'].dropna().mean()
        zero_hours = (df['Worked Hours (Number)'].fillna(0) == 0).sum()
        log_cb(f"   ↳ Avg worked hours: {avg_work:.2f} hrs  |  Zero-hour records: {zero_hours}", "info")

        log_cb("🔄 Normalising shift timings...", "info")
        df['Shift Start'] = df['Shift'].apply(get_shift_start)
        for shift, cnt in df['Shift'].value_counts().items():
            log_cb(f"   ↳ {shift}: {cnt} records", "info")

        df['Actual Shift Time'] = df['Shift'].apply(_get_actual_shift_time)

        log_cb("🕐 Calculating late login status...", "info")
        df['Late Login Status'] = df.apply(
            lambda r: calculate_late_login(r['Login Time'], r['Shift Start']), axis=1
        )
        late_count = df['Late Login Status'].apply(
            lambda x: 0 if pd.isna(x) or x in ('Early Login', 'No Delay') else 1
        ).sum()
        log_cb(f"   ↳ Late logins found: {late_count}", "info" if late_count == 0 else "warning")

        df['No of Break Counts more than 4 Times (Biometric)'] = pd.NaT
        df['No of Breaks (Biometric)'] = pd.NaT
        df['Comments'] = pd.NaT

        df.rename(columns={
            'Employee Id':    'Employee ID',
            'Team Name':      'Team/Department',
            'Shift':          'Shift Timing',
            'Idle Time':      'Idle Hours',
            'Break Time':     'Away Hours',
            'Away Time':      'Break Hours',
            'Working Hours':  'Worked Hours',
        }, inplace=True)

        log_cb("🔧 Running final adjustments (HR API, status flags, CST conversion)...", "info")
        df = final_adjustments(df, log_cb=log_cb)

        employees = df.data['Employee ID'].nunique() if hasattr(df, 'data') else '?'
        log_cb(f"✅ Timechamp processing complete — {employees} employees", "success")
        return df

    except Exception as e:
        raise Exception(f"Error processing Timechamp data: {str(e)}")


# ============================================================================
# MANUAL TIME PROCESSOR
# ============================================================================

def process_manualtime_data(df, log_cb=None):
    if log_cb is None:
        log_cb = _noop_log
    try:
        log_cb(f"📋 Loaded {len(df):,} raw records from Manual Tracking file", "info")

        df['Source'] = 'Manual Tracking'
        df['Date']   = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
        log_cb(f"   ↳ Date range: {df['Date'].min()}  →  {df['Date'].max()}", "info")

        required_cols = ['Date', 'Employee ID', 'Employee Name', 'Team/Department',
                         'Source', 'Shift Timing', 'Login Time', 'Logout Time',
                         'Manual Hours', 'Break Hours']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")
        df = df[required_cols]

        log_cb("⏱️  Converting manual hours and break hours...", "info")
        df['Worked Hours'] = df['Manual Hours'].apply(clean_time_string)
        df['Break Hours']  = df['Break Hours'].apply(clean_time_string)

        df['Worked Hours (Number)'] = df['Worked Hours'].apply(
            lambda x: round(pd.Timedelta(x).total_seconds() / 3600, 2) if x else pd.NaT)
        df['Break Hours (Number)']  = df['Break Hours'].apply(
            lambda x: round(pd.Timedelta(x).total_seconds() / 3600, 2) if x else pd.NaT)

        avg_work = df['Worked Hours (Number)'].dropna().mean()
        log_cb(f"   ↳ Average worked hours: {avg_work:.2f} hrs", "info")

        df['Shift Start']       = df['Shift Timing'].apply(get_shift_start)
        df['Actual Shift Time'] = df['Shift Timing'].apply(_get_actual_shift_time)

        log_cb("🕐 Calculating late login status...", "info")
        df['Late Login Status'] = df.apply(
            lambda r: calculate_late_login(r['Login Time'], r['Shift Start']), axis=1
        )
        late_count = df['Late Login Status'].apply(
            lambda x: 0 if pd.isna(x) or x in ('Early Login', 'No Delay') else 1
        ).sum()
        log_cb(f"   ↳ Late logins found: {late_count}", "info" if late_count == 0 else "warning")

        df['Idle Hours']  = pd.NaT
        df['Away Hours']  = pd.NaT
        df['No of Break Counts more than 4 Times (Biometric)'] = pd.NaT
        df['No of Breaks (Biometric)'] = pd.NaT
        df['Comments']    = pd.NaT

        log_cb("🔧 Running final adjustments (HR API, status flags, CST conversion)...", "info")
        df = final_adjustments(df, log_cb=log_cb)

        employees = df.data['Employee ID'].nunique() if hasattr(df, 'data') else '?'
        log_cb(f"✅ Manual tracking processing complete — {employees} employees", "success")
        return df

    except Exception as e:
        raise Exception(f"Error processing manual time data: {str(e)}")


# ============================================================================
# FILE CONSOLIDATION  (multi-file upload: merge already-processed reports)
# ============================================================================

def file_consolidate(dfs, log_cb=None):
    """
    Merge a list of already-processed DataFrames (or Styler objects) into one
    sorted, re-numbered combined report with styling applied.
    """
    if log_cb is None:
        log_cb = _noop_log
    log_cb(f"🗂️  Consolidating {len(dfs)} files...", "info")

    frames = []
    for i, item in enumerate(dfs, 1):
        raw = item.data if hasattr(item, 'data') else item
        log_cb(f"   ↳ File {i}: {len(raw):,} rows", "info")
        frames.append(raw)

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.sort_values(by=['Employee ID', 'Employee Name', 'Date'])
    if 'SL No.' in combined.columns:
        combined.drop('SL No.', axis=1, inplace=True)
    combined.insert(0, 'SL No.', range(1, len(combined) + 1))
    log_cb(f"   ↳ Combined: {len(combined):,} total rows", "info")

    log_cb("🎨 Applying styles to consolidated report...", "info")
    styled = combined.style.apply(break_highlight_status, axis=1)
    styled = styled.map(Attendance_highlight_status, subset=['Attendance'])
    styled = styled.map(Date_highlight_status,       subset=['Date'])
    return styled


# ============================================================================
# PIVOT TABLE
# ============================================================================

def create_pivot_table(df, log_cb=None):
    if log_cb is None:
        log_cb = _noop_log
    log_cb("📊 Building Overall Summary pivot table...", "info")

    pivot = df.data.copy() if hasattr(df, 'data') else df.copy()

    attendance_mapping = {'Absent': 0, 'WeekOff': 0, 'Full day': 1, 'Half day': 0.5}
    pivot['No of Worked Days'] = pivot['Attendance'].map(attendance_mapping)

    pivot['No of Worked below 8 hrs'] = pivot['Worked Hours Status'].apply(lambda x: 1 if x == 'Worked below 8 hrs' else 0)
    pivot['No of Worked below 4 hrs'] = pivot['Worked Hours Status'].apply(lambda x: 1 if x == 'Worked below 4 hrs' else 0)
    pivot['No of More Break Hours']   = pivot['Break Hours Status'].apply(lambda x: 1 if isinstance(x, str) else 0)
    pivot['No of Attendance Mismatch'] = pivot['Attendance Mismatch'].apply(lambda x: 1 if isinstance(x, str) else 0)

    half_leaves = ['Sick Leave - Half', 'EL HD', 'Leave - Half Day', 'Absent - Half Day']

    pivot['Total HR Attendance Hours']    = pivot['HR Attendance'].apply(
        lambda v: 0.5 * 8 if v in half_leaves else (1 * 8 if v == 'Present' else 0))
    pivot['No of Worked Days (Attendance)'] = pivot['HR Attendance'].apply(
        lambda v: 0.5 if v in half_leaves else (1 if v == 'Present' else 0))

    pivot['No of late login']    = pivot['Late Login Status'].apply(map_late_login_status)
    pivot['No of Early Logout']  = pivot['Early Logout Status'].apply(
        lambda x: 0 if pd.isna(x) or x == 'No Early Logout' else 1)

    pivot['No of Breaks (Biometric)'] = pivot['No of Breaks (Biometric)'].replace('', 0).fillna(0).astype(int)
    pivot['No of Break Counts more than 4 Times (Biometric)'] = pivot['No of Breaks (Biometric)'].apply(
        lambda x: 1 if x > 4 else 0)

    pivot['Total Worked Hours'] = pivot['Worked Hours'].apply(parse_time_string)
    pivot['Total Break Hours']  = pivot['Break Hours'].apply(parse_time_string)

    # Fix: pass through already-timedelta values without unit='D' corruption
    for col in ['Total Worked Hours', 'Total Break Hours']:
        pivot[col] = pd.to_timedelta(pivot[col], errors='coerce')

    pivot = pivot.groupby(['Employee ID', 'Employee Name', 'Team/Department', 'Source']).agg({
        'No of Worked Days': 'sum',
        'No of Worked Days (Attendance)': 'sum',
        'No of late login': 'sum',
        'No of Early Logout': 'sum',
        'No of Worked below 4 hrs': 'sum',
        'No of Worked below 8 hrs': 'sum',
        'No of More Break Hours': 'sum',
        'No of Break Counts more than 4 Times (Biometric)': 'sum',
        'Total HR Attendance Hours': 'sum',
        'Total Worked Hours': 'sum',
        'Total Break Hours': 'sum',
        'No of Attendance Mismatch': 'sum',
    }).fillna(0).reset_index()

    pivot['Total HR Attendance Hours'] = pd.to_timedelta(pivot['Total HR Attendance Hours'], unit='h', errors='coerce')
    pivot['Worked Hours Diff']   = pivot['Total Worked Hours'] - pivot['Total HR Attendance Hours']
    pivot['Worked Hours Diff']   = pivot['Worked Hours Diff'].apply(
        lambda x: round(x.total_seconds() / 3600, 2) if pd.notna(x) else pd.NaT)
    pivot['Total Worked Hours']  = pivot['Total Worked Hours'].apply(
        lambda x: round(x.total_seconds() / 3600, 2) if pd.notna(x) else pd.NaT)
    pivot['Total Break Hours']   = pivot['Total Break Hours'].apply(
        lambda x: round(x.total_seconds() / 3600, 2) if pd.notna(x) else pd.NaT)
    pivot['Total HR Attendance Hours'] = pivot['Total HR Attendance Hours'].apply(
        lambda x: round(x.total_seconds() / 3600, 2) if pd.notna(x) else pd.NaT)

    pivot.insert(0, 'SL No.', range(1, len(pivot) + 1))
    log_cb(f"   ↳ Pivot table: {len(pivot)} employee rows", "info")
    return pivot


# ============================================================================
# HIGHLIGHTED DATA (Issues-only view)
# ============================================================================

def create_highlighted_data(df, log_cb=None):
    if log_cb is None:
        log_cb = _noop_log
    log_cb("🔦 Building Highlighted Users sheet (issues only)...", "info")

    data = df.data.copy() if hasattr(df, 'data') else df.copy()
    data = data.replace({'Early Login': pd.NaT, 'No Delay': pd.NaT, 'No Early Logout': pd.NaT})
    data = data.dropna(
        subset=['Late Login Status', 'Early Logout Status', 'Worked Hours Status',
                'Break Hours Status', 'No of Break Counts more than 4 Times (Biometric)',
                'Attendance Mismatch'],
        how='all',
    )
    if 'SL No.' in data.columns:
        data.drop('SL No.', axis=1, inplace=True)
    data.insert(0, 'SL No.', range(1, len(data) + 1))

    log_cb(f"   ↳ {len(data)} flagged records across all issue types", "info")

    styled = data.style.apply(break_highlight_status, axis=1)
    styled = styled.map(Attendance_highlight_status, subset=['Attendance'])
    styled = styled.map(Date_highlight_status,       subset=['Date'])
    return styled


# ============================================================================
# EXCEL OUTPUT  (matches original script: centre-align, Verdana, dotted borders,
#                colour-coded tabs, per-column alignment)
# ============================================================================

def save_to_excel(df, pivot_table, highlight_data, source_type, log_cb=None):
    if log_cb is None:
        log_cb = _noop_log
    try:
        temp_dir = tempfile.mkdtemp()

        df_copy = df.data.copy() if hasattr(df, 'data') else df.copy()
        if 'Date' in df_copy.columns:
            df_copy['Date'] = pd.to_datetime(df_copy['Date'])
            min_date = df_copy['Date'].min()
            max_date = df_copy['Date'].max()
            min_str  = min_date.strftime('%d-%b-%Y')
            max_str  = max_date.strftime('%d-%b-%Y')
            file_name = (f"{source_type} Report {min_str} to {max_str}.xlsx"
                         if min_str != max_str else
                         f"{source_type} Report {min_str}.xlsx")
        else:
            file_name = f"{source_type} Report.xlsx"

        file_path = os.path.join(temp_dir, file_name)
        log_cb(f"💾 Writing Excel file: {file_name}", "info")

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer,             index=False, sheet_name='Detailed Report')
            highlight_data.to_excel(writer, index=False, sheet_name='Highlighted Users')
            pivot_table.to_excel(writer,    index=False, sheet_name='Overall Summary')

            workbook = writer.book

            # Columns that should be centre-aligned (matches original script exactly)
            center_cols = {
                'No of Worked Days', 'No of Attendance Mismatch', 'No of Worked Days (Attendance)',
                'No of Break Hours more than 1 Hours', 'No of late login', 'No of Early Logout',
                'Total HR Attendance Hours', 'No of Break Counts more than 4 Times (Biometric)',
                'SL No.', 'Login Time', 'Logout Time', 'No of Breaks (Biometric)',
                'Attendance', 'Day', 'Worked Hours (Number)', 'Break Hours (Number)',
                'No of Worked below 4 hrs', 'No of Worked below 8 hrs',
                'No of More Break Hours', 'Away Hours',
            }

            tab_colors = {
                'Detailed Report':  'CCFFCC',
                'Highlighted Users': 'FFCCCC',
                'Overall Summary':  'FFD700',
            }

            side         = Side(border_style='dotted', color='000000')
            border_style = Border(left=side, right=side, top=side, bottom=side)

            for sheet_name in ['Detailed Report', 'Highlighted Users', 'Overall Summary']:
                ws = workbook[sheet_name]
                ws.sheet_properties.tabColor = tab_colors[sheet_name]

                # Header row
                for cell in ws[1]:
                    cell.font      = Font(name='Verdana', size=9, bold=True)
                    cell.fill      = PatternFill(start_color='9bc2e6', end_color='9bc2e6', fill_type='solid')
                    cell.alignment = Alignment(vertical='bottom', horizontal='left')
                    cell.border    = border_style

                # Identify centre-aligned column indices from header values
                center_indices = {
                    ws.cell(row=1, column=i).column
                    for i in range(1, ws.max_column + 1)
                    if ws.cell(row=1, column=i).value in center_cols
                }

                total_rows = ws.max_row - 1
                log_cb(f"   ↳ Styling '{sheet_name}' — {total_rows:,} rows...", "info")

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                        min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.font   = Font(name='Verdana', size=9, color='000000')
                        cell.border = border_style
                        cell.alignment = Alignment(
                            vertical='center' if cell.column in center_indices else 'bottom',
                            horizontal='center' if cell.column in center_indices else 'left',
                        )

        log_cb(f"✅ Excel file saved — {file_name}", "success")
        return file_path, file_name

    except Exception as e:
        raise Exception(f"Error saving Excel file: {str(e)}")
